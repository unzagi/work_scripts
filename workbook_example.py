import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\sbrooks\Downloads\RM CPE Router Specification and Build v10.xlsx")
print("Workbook ", wb, " is loaded")

#Manual Imput
#print("Please enter the sheet name")

#Auto Override
sheet = wb.get_sheet_by_name("DESIGN TEMPLATES")


##print("Please enter the customer RMID: ")
##rmidChoice = input()
##fileName = "csr01.id" + str(rmidChoice) + "template.details.txt"
##
##
##print("Please enter the Juniper Template Number:")
##myChoice = input()
##print(myChoice)

cell1=[]
cell2=[]
cell3=[]
cell4=[]
cell5=[]
cell6=[]
cell7=[]
cell8=[]
cell9=[]
cell10=[]
cell11=[]
cell12=[]
cell13=[]
cell14=[]

##print("Please enter the Juniper Template Number:")
##templateChoice = input()


for i in range(1,sheet.max_row + 1):
    cell1.append(sheet.cell(row=i, column=1).value)
    cell2.append(sheet.cell(row=i, column=2).value)
    cell3.append(sheet.cell(row=i, column=3).value)
    cell4.append(sheet.cell(row=i, column=4).value)
    cell5.append(sheet.cell(row=i, column=5).value)
    cell6.append(sheet.cell(row=i, column=6).value)
    cell7.append(sheet.cell(row=i, column=7).value)
    cell8.append(sheet.cell(row=i, column=8).value)
    cell9.append(sheet.cell(row=i, column=9).value)
    cell10.append(sheet.cell(row=i, column=10).value)
    cell11.append(sheet.cell(row=i, column=11).value)
    cell12.append(sheet.cell(row=i, column=12).value)
    cell13.append(sheet.cell(row=i, column=13).value)
    cell14.append(sheet.cell(row=i, column=14).value)

##for i in range(1,sheet.max_row):
##        print(cell1[i])
##        print(cell2[i])
##        print(cell3[i])
##        print(cell4[i])
##        print(cell5[i])
##        print(cell6[i])
##        print(cell7[i])
##        print(cell8[i])
##        print(cell9[i])
##        print(cell10[i])
##        print(cell11[i])
##        print(cell12[i])
##        print(cell13[i])
##        print(cell14[i])

        
for j in range(1,sheet.max_row):
    if cell5[j] == 'Juniper':
        if cell3[j] == 'Y':
            sheet = (cell2[j],cell3[j],cell5[j],cell6[j],cell7[j],cell10[j],cell11[j])            
print(sheet)
##            if cell2[j] == str(templateChoice):
##                print("Cell: \n",cell2[j])
##                print("Choice: \n",templateChoice)
##                i = sheet.max_row
                
##print("Template Choice is: ", templateChoice)                
##print(j,cell2[j],cell3[j],cell5[j],cell6[j],cell7[j],cell10[j],cell11[j])



##                text_file = open((fileName), "w")
##                text_file.write(output)
##                text_file.close()
##                print(fileName,"has been written to disk")


       
##        if cell3[i] == 'Y':
##            if myChoice in cell7:
##            print(i,cell2[i],cell3[i],cell5[i],cell6[i],cell7[i],cell10[i],cell11[i])
##            print(i,cell1,cell2,cell3,cell4,cell5,cell6,cell7,cell8,cell9,cell10,cell11,cell2,cell13,cell14)
            
##            print("Please enter the Juniper Template Number:")
##            templateChoice = input()
##            print(templateChoice)
##            if cell2 == templateChoice: 
##                text_file = open((fileName), "w")
##                text_file.write(output)
##                text_file.close()
##                print(fileName,"has been written to disk")



##for i in range(1, sheet.max_row + 1):
##    cell1 = sheet.cell(row=i, column=1).value
##    cell2 = sheet.cell(row=i, column=2).value
##    cell3 = sheet.cell(row=i, column=3).value
##    cell4 = sheet.cell(row=i, column=4).value
##    cell5 = sheet.cell(row=i, column=5).value
##    cell6 = sheet.cell(row=i, column=6).value
##    cell7 = sheet.cell(row=i, column=7).value
##    cell8 = sheet.cell(row=i, column=8).value
##    cell9 = sheet.cell(row=i, column=9).value
##    cell10 = sheet.cell(row=i, column=10).value
##    cell11 = sheet.cell(row=i, column=11).value
##    cell12 = sheet.cell(row=i, column=12).value
##    cell13 = sheet.cell(row=i, column=13).value
##    cell14 = sheet.cell(row=i, column=14).value
##    print (i)

