import openpyxl


##def(createCells):
##    for i in range(10,12 + 1):
##        print("cell" + i + " = sheet.cell(row=i, column=" + i + ").value")


wb = openpyxl.load_workbook(r"C:\Users\sbrooks\Downloads\SxGfL VirginMedia Connections.xlsx")
print("Workbook ", wb, " is loaded")

#Manual Imput
#print("Please enter the sheet name")
#myChoice = input()
#print(myChoice)


#Auto Override
sheet = wb.get_sheet_by_name("Data")

print("Please enter the customer ID: ")
myChoice = input()
str(myChoice)
print(myChoice)

for i in range(3, sheet.max_row + 1):
    cell1 = sheet.cell(row=i, column=1).value
    cell2 = sheet.cell(row=i, column=2).value
    cell3 = sheet.cell(row=i, column=3).value
    cell4 = sheet.cell(row=i, column=4).value
    cell5 = sheet.cell(row=i, column=5).value
    cell6 = sheet.cell(row=i, column=6).value
    cell7 = sheet.cell(row=i, column=7).value
    cell8 = sheet.cell(row=i, column=8).value
    cell9 = sheet.cell(row=i, column=9).value
    cell10 = sheet.cell(row=i, column=10).value
    cell11 = sheet.cell(row=i, column=11).value
    cell12 = sheet.cell(row=i, column=12).value

    if str(myChoice) in str(cell2):
        print(i,cell1,cell2,cell3,cell4,cell5,cell6,cell7,cell8,cell8,cell9,cell10,cell11,cell12)
        

   
    

        

        

